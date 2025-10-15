import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from bcb import PTAX
import io
import logging

# ====== Configuração de Logging ======
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ====== Constantes ======
REQUIRED_COLUMNS = [
    "Tipo de dívida",
    "Situação da dívida",
    "Valor a liberar ou assumir (na moeda de contratação)",
    "Moeda da contratação, emissão ou assunção"
]

MAPA_MOEDAS = {
    "Real": "BRL",
    "Dólar dos EUA": "USD",
    "Euro": "EUR",
    "Direito Especial - SDR": "XDR",
    "Iene": "JPY"
}

SIMBOLOS_MOEDAS = {
    "Real": "R$",
    "Dólar dos EUA": "US$",
    "Euro": "€",
    "Direito Especial - SDR": "SDR",
    "Iene": "¥"
}

# Intervalos bimestrais do RREO: (início período, fim período, data referência cotação)
# Formato: ((mês_ini, dia_ini), (mês_fim, dia_fim), (mês_ref, dia_ref))
INTERVALOS_RREO = [
    ((3, 31), (5, 30), (2, 28)),   # Mar/Abr → Cotação 28/fev
    ((5, 31), (7, 30), (4, 30)),   # Mai/Jun → Cotação 30/abr
    ((7, 31), (9, 30), (6, 30)),   # Jul/Ago → Cotação 30/jun
    ((10, 1), (11, 30), (8, 31)),  # Set/Out → Cotação 31/ago
    ((12, 1), (1, 30), (10, 31)),  # Nov/Dez → Cotação 31/out
    ((1, 31), (3, 30), (12, 31))   # Jan/Fev → Cotação 31/dez (ano anterior)
]

MAX_ARQUIVO_MB = 50

# ====== Inicializa PTAX ======
ptax = PTAX()
ep_cotacao = ptax.get_endpoint('CotacaoMoedaDia')

# ====== Funções Auxiliares ======
def formatar_numero_brasil(valor, casas_decimais=2):
    """
    Formata número no padrão brasileiro: ponto para milhar, vírgula para decimal.
    
    Args:
        valor: Número a ser formatado
        casas_decimais: Quantidade de casas decimais (padrão: 2)
    
    Returns:
        String formatada no padrão brasileiro
    """
    if not isinstance(valor, (int, float)):
        return valor
    
    # Formata com casas decimais e separadores
    formatado = f"{valor:,.{casas_decimais}f}"
    # Converte para padrão brasileiro
    formatado = formatado.replace(',', 'TEMP').replace('.', ',').replace('TEMP', '.')
    return formatado


def validar_csv(df):
    """
    Valida se o CSV possui as colunas necessárias.
    
    Args:
        df: DataFrame a ser validado
    
    Raises:
        ValueError: Se colunas obrigatórias estiverem ausentes
    """
    # Normaliza nomes das colunas
    df.columns = [c.strip() for c in df.columns]
    
    # Verifica colunas ausentes
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        logger.error(f"Colunas ausentes no CSV: {missing}")
        raise ValueError(f"O arquivo CSV não possui as colunas obrigatórias: {', '.join(missing)}")
    
    logger.info("Validação do CSV concluída com sucesso")
    return True


def data_cotacao():
    """
    Determina a data de referência para cotação baseada nos períodos do RREO.
    Retorna o último dia útil do bimestre anterior.
    
    Returns:
        String com data no formato MM/DD/YYYY
    """
    hoje = datetime.today()
    ano = hoje.year
    
    # Busca o intervalo correspondente
    for inicio, fim, referencia in INTERVALOS_RREO:
        mes_ini, dia_ini = inicio
        mes_fim, dia_fim = fim
        mes_ref, dia_ref = referencia
        
        # Ajusta o ano para intervalos que cruzam o ano
        if mes_ini < mes_fim:
            ini = datetime(ano, mes_ini, dia_ini)
            fim_periodo = datetime(ano, mes_fim, dia_fim)
        else:
            ini = datetime(ano, mes_ini, dia_ini)
            fim_periodo = datetime(ano + 1, mes_fim, dia_fim)
        
        # Verifica se a data atual está no intervalo
        if ini <= hoje <= fim_periodo:
            # Dezembro do ano anterior para o intervalo Jan/Fev
            ano_ref = ano if mes_ref != 12 else ano - 1
            data_base = datetime(ano_ref, mes_ref, dia_ref)
            break
    else:
        # Fallback: último dia de fevereiro
        logger.warning("Data atual fora dos intervalos definidos, usando fallback")
        data_base = datetime(ano, 2, 28)
    
    # Ajusta para dia útil (não final de semana)
    while data_base.weekday() >= 5:  # 5=Sábado, 6=Domingo
        data_base -= timedelta(days=1)
        logger.info(f"Ajustando para dia útil: {data_base.strftime('%d/%m/%Y')}")
    
    logger.info(f"Data de referência calculada: {data_base.strftime('%d/%m/%Y')}")
    return data_base.strftime("%m/%d/%Y")


@st.cache_data(ttl=28800)  # Cache por 8 horas
def cotacao_bacen(moeda, data_ref):
    """
    Busca cotação PTAX de venda no Banco Central para uma moeda e data específicas.
    Tenta até 5 dias úteis anteriores se não encontrar na data informada.
    
    Args:
        moeda: Código da moeda (USD, EUR, etc.)
        data_ref: Data de referência no formato MM/DD/YYYY
    
    Returns:
        Tupla (cotação, data_utilizada) ou ("-", "-") se não encontrar
    """
    if moeda == "BRL":
        logger.info("Moeda BRL, retornando cotação 1.0")
        return 1.0, ""
    
    # Para SDR, não busca cotação na API
    if moeda == "XDR":
        logger.info("Moeda SDR, retornando 'Sem cotação'")
        return "Sem cotação", "-"
    
    # Tenta buscar cotação nos últimos 5 dias úteis
    for i in range(5):
        dt_busca = (datetime.strptime(data_ref, "%m/%d/%Y") - timedelta(days=i)).strftime("%m/%d/%Y")
        
        try:
            logger.info(f"Buscando cotação {moeda} para {dt_busca}")
            df = ep_cotacao.query().parameters(moeda=moeda, dataCotacao=dt_busca).collect()
            
            if not df.empty:
                # Filtra apenas fechamento PTAX
                fechamento = df[df['tipoBoletim'] == 'Fechamento PTAX']
                
                if not fechamento.empty:
                    # Cotação de venda: 'cotacaoVenda'
                    cotacao = float(fechamento['cotacaoVenda'].values[-1])
                    data_formatada = datetime.strptime(dt_busca, "%m/%d/%Y").strftime("%d/%m/%Y")
                    logger.info(f"Cotação encontrada: {cotacao} em {data_formatada}")
                    return cotacao, data_formatada
                    
        except Exception as e:
            logger.warning(f"Erro ao buscar cotação {moeda} para {dt_busca}: {e}")
            continue
    
    logger.error(f"Não foi possível obter cotação para {moeda}")
    return "-", "-"


def processar_csv(df_csv):
    """
    Processa o CSV de dívidas, filtra registros relevantes, agrupa por moeda
    e converte valores para BRL usando cotações do Banco Central.
    
    Args:
        df_csv: DataFrame com dados do CSV
    
    Returns:
        DataFrame com resumo processado ou None se não houver registros
    """
    logger.info("Iniciando processamento do CSV")
    
    # Normaliza nomes das colunas
    df_csv.columns = [c.strip() for c in df_csv.columns]
    
    # Filtros aplicados
    logger.info("Aplicando filtros: tipo=empréstimo, situação=vigente, valor>0")
    df_filtrado = df_csv[
        (df_csv["Tipo de dívida"].str.strip().str.lower() == "empréstimo ou financiamento") &
        (df_csv["Situação da dívida"].str.strip().str.lower() == "vigente") &
        (df_csv["Valor a liberar ou assumir (na moeda de contratação)"] > 0)
    ]
    
    if df_filtrado.empty:
        logger.warning("Nenhum registro encontrado após aplicar filtros")
        return None
    
    logger.info(f"Registros encontrados após filtros: {len(df_filtrado)}")
    
    # Agrupa por moeda
    resumo = (
        df_filtrado.groupby("Moeda da contratação, emissão ou assunção")
        ["Valor a liberar ou assumir (na moeda de contratação)"]
        .sum()
        .reset_index()
    )
    
    logger.info(f"Moedas encontradas: {resumo['Moeda da contratação, emissão ou assunção'].tolist()}")
    
    # Obtém data de referência para cotação
    data_ref = data_cotacao()
    
    # Processa cada moeda
    valores = []
    for _, row in resumo.iterrows():
        nome_moeda = row["Moeda da contratação, emissão ou assunção"]
        valor = row["Valor a liberar ou assumir (na moeda de contratação)"]
        
        codigo = MAPA_MOEDAS.get(nome_moeda)
        if not codigo:
            logger.warning(f"Moeda não mapeada: {nome_moeda}")
            continue
        
        # Busca cotação
        cot, data_usada = cotacao_bacen(codigo, data_ref)
        
        # Para SDR, não calcula valor em BRL
        if nome_moeda == "Direito Especial - SDR":
            valor_brl = "-"
        else:
            # Calcula valor em BRL para outras moedas
            valor_brl = valor * cot if isinstance(cot, float) else valor
        
        valores.append((nome_moeda, valor, cot, data_usada, valor_brl))
        
        if nome_moeda == "Direito Especial - SDR":
            logger.info(f"SDR: {valor} × Sem cotação = -")
        else:
            logger.info(f"{nome_moeda}: {valor} × {cot} = R$ {valor_brl:,.2f}")
    
    # Cria DataFrame de saída
    df_saida = pd.DataFrame(
        valores, 
        columns=["Moeda", "Valor a Liberar", "Cotação", "Data da Cotação", "Valor em BRL"]
    )
    
    # Adiciona linha TOTAL (considerando apenas moedas com valor em BRL)
    total_brl = df_saida[df_saida["Valor em BRL"] != "-"]["Valor em BRL"].sum()
    total = pd.DataFrame(
        [["TOTAL", "", "", "", total_brl]], 
        columns=df_saida.columns
    )
    df_saida = pd.concat([df_saida, total], ignore_index=True)
    
    logger.info(f"Processamento concluído. Total em BRL: R$ {total_brl:,.2f}")
    return df_saida


def extrair_registros_detalhados(df_csv):
    """
    Extrai os registros que atendem aos critérios com colunas específicas.
    
    Args:
        df_csv: DataFrame com dados do CSV original
    
    Returns:
        DataFrame com registros detalhados ou None se não houver registros
    """
    logger.info("Extraindo registros detalhados")
    
    # Normaliza nomes das colunas
    df_csv.columns = [c.strip() for c in df_csv.columns]
    
    # Aplica os mesmos filtros
    df_filtrado = df_csv[
        (df_csv["Tipo de dívida"].str.strip().str.lower() == "empréstimo ou financiamento") &
        (df_csv["Situação da dívida"].str.strip().str.lower() == "vigente") &
        (df_csv["Valor a liberar ou assumir (na moeda de contratação)"] > 0)
    ].copy()
    
    if df_filtrado.empty:
        return None
    
    # Seleciona apenas as colunas desejadas (índices de coluna começam em 0)
    # B=1, C=2, I=8, L=11, M=12, N=13, S=18, AF=31
    colunas_interesse = df_csv.columns[[1, 2, 8, 11, 12, 13, 18, 31]].tolist()
    
    df_detalhado = df_filtrado[colunas_interesse].copy()
    
    # Renomeia as colunas para os nomes desejados
    df_detalhado.columns = ["UF", "Ente", "Nome do Credor", "Moeda", "Valor contratado", "Taxa de juros", "Valor a liberar", "Data da quitação"]
    
    logger.info(f"Extraídos {len(df_detalhado)} registros detalhados")
    return df_detalhado


def formatar_para_exibicao(df_resumo):
    """
    Formata o DataFrame para exibição com símbolos de moeda e padrão brasileiro.
    
    Args:
        df_resumo: DataFrame com dados processados
    
    Returns:
        DataFrame formatado para exibição
    """
    df_vis = df_resumo.copy()
    
    for i, row in df_vis.iterrows():
        moeda = row["Moeda"]
        
        if moeda != "TOTAL":
            # Formata valor a liberar com símbolo da moeda
            simbolo = SIMBOLOS_MOEDAS.get(moeda, "")
            valor_formatado = formatar_numero_brasil(row["Valor a Liberar"], 2)
            df_vis.at[i, "Valor a Liberar"] = f"{simbolo} {valor_formatado}"
            
            # Formata cotação
            if isinstance(row["Cotação"], (int, float)):
                df_vis.at[i, "Cotação"] = formatar_numero_brasil(row["Cotação"], 5)
            elif row["Cotação"] == "Sem cotação":
                df_vis.at[i, "Cotação"] = "Sem cotação"
            
            # Substitui vazio por "-"
            if df_vis.at[i, "Data da Cotação"] == "":
                df_vis.at[i, "Data da Cotação"] = "-"
        else:
            # Linha TOTAL
            df_vis.at[i, "Valor a Liberar"] = "-"
            df_vis.at[i, "Cotação"] = "-"
            df_vis.at[i, "Data da Cotação"] = "-"
        
        # Formata valor em BRL (todas as linhas)
        if row["Valor em BRL"] != "-":
            valor_brl_formatado = formatar_numero_brasil(row["Valor em BRL"], 2)
            df_vis.at[i, "Valor em BRL"] = f"R$ {valor_brl_formatado}"
        else:
            df_vis.at[i, "Valor em BRL"] = "-"
    
    return df_vis


def formatar_detalhes_para_exibicao(df_detalhes):
    """
    Formata o DataFrame de detalhes para exibição.
    
    Args:
        df_detalhes: DataFrame com registros detalhados
    
    Returns:
        DataFrame formatado para exibição
    """
    df_vis = df_detalhes.copy()
    
    # Formata valores numéricos
    for i, row in df_vis.iterrows():
        # Formata valor a liberar
        if pd.notna(row["Valor a liberar"]) and isinstance(row["Valor a liberar"], (int, float)):
            df_vis.at[i, "Valor a liberar"] = formatar_numero_brasil(row["Valor a liberar"], 2)
        
        # Formata valor contratado (máscara de moeda sem símbolo)
        if pd.notna(row["Valor contratado"]) and isinstance(row["Valor contratado"], (int, float)):
            df_vis.at[i, "Valor contratado"] = formatar_numero_brasil(row["Valor contratado"], 2)
        
        # Formata taxa de juros (se for numérico)
        if pd.notna(row["Taxa de juros"]) and isinstance(row["Taxa de juros"], (int, float)):
            df_vis.at[i, "Taxa de juros"] = formatar_numero_brasil(row["Taxa de juros"], 2)
        
        # Formata data da quitação (se for datetime)
        if pd.notna(row["Data da quitação"]):
            if isinstance(row["Data da quitação"], (datetime, pd.Timestamp)):
                df_vis.at[i, "Data da quitação"] = row["Data da quitação"].strftime("%d/%m/%Y")
            elif isinstance(row["Data da quitação"], str):
                # Tenta converter string para data
                try:
                    data_obj = datetime.strptime(row["Data da quitação"], "%Y-%m-%d")
                    df_vis.at[i, "Data da quitação"] = data_obj.strftime("%d/%m/%Y")
                except:
                    # Mantém o valor original se não conseguir converter
                    pass
    
    return df_vis


def gerar_html_tabela(df_vis):
    """
    Gera HTML customizado para exibir a tabela com formatação especial.
    
    Args:
        df_vis: DataFrame formatado para exibição
    
    Returns:
        String com HTML da tabela
    """
    html = """
    <style>
    .dataframe-custom {
        width: 100%;
        border-collapse: collapse;
        margin: 20px 0;
        font-size: 14px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .dataframe-custom th {
        background-color: #1f77b4;
        color: white;
        padding: 12px;
        text-align: left;
        font-weight: 600;
    }
    .dataframe-custom td {
        padding: 10px 12px;
        border-bottom: 1px solid #ddd;
    }
    .dataframe-custom tr:hover:not(.total-row) {
        background-color: #f5f5f5;
    }
    .total-row {
        font-weight: bold;
        background-color: rgba(31, 119, 180, 0.15) !important;
        border-top: 2px solid #1f77b4;
    }
    .total-row td {
        padding: 12px;
        font-size: 15px;
        color: inherit;
    }
    </style>
    <table class="dataframe-custom">
    <thead>
        <tr>
            <th>Moeda</th>
            <th>Valor a Liberar</th>
            <th>Cotação</th>
            <th>Data da Cotação</th>
            <th>Valor em BRL</th>
        </tr>
    </thead>
    <tbody>
    """
    
    for _, row in df_vis.iterrows():
        classe = 'class="total-row"' if row["Moeda"] == "TOTAL" else ''
        html += f'<tr {classe}>'
        html += f'<td>{row["Moeda"]}</td>'
        html += f'<td>{row["Valor a Liberar"]}</td>'
        html += f'<td>{row["Cotação"]}</td>'
        html += f'<td>{row["Data da Cotação"]}</td>'
        html += f'<td>{row["Valor em BRL"]}</td>'
        html += '</tr>'
    
    html += "</tbody></table>"
    return html


def gerar_html_tabela_detalhes(df_vis):
    """
    Gera HTML customizado para exibir a tabela de detalhes.
    
    Args:
        df_vis: DataFrame formatado para exibição
    
    Returns:
        String com HTML da tabela
    """
    html = """
    <style>
    .dataframe-detalhes {
        width: 100%;
        border-collapse: collapse;
        margin: 20px 0;
        font-size: 13px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .dataframe-detalhes th {
        background-color: #2c3e50;
        color: white;
        padding: 12px;
        text-align: left;
        font-weight: 600;
    }
    .dataframe-detalhes td {
        padding: 10px 12px;
        border-bottom: 1px solid #ddd;
    }
    .dataframe-detalhes tr:hover {
        background-color: rgba(0, 0, 0, 0.05);
    }
    .dataframe-detalhes tr:nth-child(even) {
        background-color: rgba(0, 0, 0, 0.02);
    }
    
    /* Ajustes para modo escuro */
    @media (prefers-color-scheme: dark) {
        .dataframe-detalhes {
            box-shadow: 0 2px 4px rgba(255,255,255,0.1);
        }
        .dataframe-detalhes td {
            border-bottom: 1px solid #444;
        }
        .dataframe-detalhes tr:hover {
            background-color: rgba(255, 255, 255, 0.1);
        }
        .dataframe-detalhes tr:nth-child(even) {
            background-color: rgba(255, 255, 255, 0.05);
        }
    }
    </style>
    <table class="dataframe-detalhes">
    <thead>
        <tr>
    """
    
    # Cabeçalhos
    for col in df_vis.columns:
        html += f'<th>{col}</th>'
    
    html += """
        </tr>
    </thead>
    <tbody>
    """
    
    # Dados
    for _, row in df_vis.iterrows():
        html += '<tr>'
        for col in df_vis.columns:
            html += f'<td>{row[col]}</td>'
        html += '</tr>'
    
    html += "</tbody></table>"
    return html


def gerar_excel_completo(df_csv_original, df_resumo):
    """
    Gera arquivo Excel com todos os dados do CSV original e tabela de resumo.
    Pinta as linhas que atendem aos critérios com cores diferentes por moeda.
    Posiciona a tabela de resumo alinhando "Valor a Liberar" com a coluna correspondente do CSV.
    Oculta colunas específicas e ajusta largura das colunas conforme especificado.
    
    Args:
        df_csv_original: DataFrame com TODOS os dados do CSV original
        df_resumo: DataFrame com a tabela de resumo
    
    Returns:
        BytesIO com o arquivo Excel gerado
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    logger.info("Gerando arquivo Excel completo")
    
    # Cores por moeda (tons pastéis para melhor leitura)
    CORES_MOEDAS = {
        "Real": "C6EFCE",           # Verde claro
        "Dólar dos EUA": "FFF2CC",  # Amarelo claro
        "Euro": "DDEBF7",           # Azul claro
        "Direito Especial - SDR": "FCE4D6",  # Laranja claro
        "Iene": "E2EFDA"            # Verde água claro
    }
    
    # Colunas para ocultar: A, D, E, G, H, J, K, O, W até AE (inclusive)
    colunas_para_ocultar = ['A', 'D', 'E', 'G', 'H', 'J', 'K', 'O'] + [get_column_letter(i) for i in range(23, 32)]  # W=23 até AE=31
    
    buffer = io.BytesIO()
    
    # Cria o Excel com openpyxl
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Escreve dados originais
        df_csv_original.to_excel(writer, sheet_name='Dados', index=False, startrow=0)
        
        # Pega a planilha
        workbook = writer.book
        worksheet = writer.sheets['Dados']
        
        # Congela a primeira linha (cabeçalho) para facilitar a visualização com filtros
        worksheet.freeze_panes = "A2"

        # Formata cabeçalho dos dados originais
        header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplica formatação no cabeçalho
        for col_num, column in enumerate(df_csv_original.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # ====== OCULTA COLUNAS ESPECÍFICAS ======
        for col_letter in colunas_para_ocultar:
            worksheet.column_dimensions[col_letter].hidden = True
        
        # ====== AJUSTA LARGURA DAS COLUNAS ESPECÍFICAS ======
        # Coluna I: largura para caber "Caixa Econômica Federal" (aproximadamente 25 caracteres)
        worksheet.column_dimensions['I'].width = 25
        
        # Coluna L: largura para caber "Direito Especial - SDR" (aproximadamente 22 caracteres)
        worksheet.column_dimensions['L'].width = 22
        
        # Coluna M: largura para caber "Valor da contratação, em" (aproximadamente 24 caracteres)
        worksheet.column_dimensions['M'].width = 24
        
        # ====== ENCONTRA A COLUNA "Valor a liberar ou assumir (na moeda de contratação)" ======
        coluna_valor_liberar = None
        for col_num, column in enumerate(df_csv_original.columns, 1):
            if column.strip() == "Valor a liberar ou assumir (na moeda de contratação)":
                coluna_valor_liberar = col_num
                break
        
        # Se não encontrou a coluna, usa a coluna 2 como fallback
        if coluna_valor_liberar is None:
            logger.warning("Coluna 'Valor a liberar ou assumir (na moeda de contratação)' não encontrada, usando coluna B como fallback")
            coluna_valor_liberar = 2
        
        # ====== FORMATA COLUNA DE VALOR DOS DADOS ORIGINAIS ======
        # Formata todos os valores da coluna "Valor a liberar ou assumir (na moeda de contratação)" com máscara de moeda
        for idx in range(len(df_csv_original)):
            excel_row = idx + 2  # +2 porque Excel começa em 1 e tem cabeçalho
            cell_valor = worksheet.cell(row=excel_row, column=coluna_valor_liberar)
            # Aplica formatação de moeda genérica (sem símbolo específico)
            cell_valor.number_format = '#,##0.00'
        
        # Identifica registros que atendem aos critérios e pinta por moeda
        for idx, row in df_csv_original.iterrows():
            excel_row = idx + 2  # +2 porque Excel começa em 1 e tem cabeçalho
            
            # Verifica se atende aos critérios
            atende_criterios = (
                str(row.get("Tipo de dívida", "")).strip().lower() == "empréstimo ou financiamento" and
                str(row.get("Situação da dívida", "")).strip().lower() == "vigente" and
                pd.notna(row.get("Valor a liberar ou assumir (na moeda de contratação)")) and
                float(row.get("Valor a liberar ou assumir (na moeda de contratação)", 0)) > 0
            )
            
            if atende_criterios:
                moeda = row.get("Moeda da contratação, emissão ou assunção", "")
                cor = CORES_MOEDAS.get(moeda, "FFFFFF")
                fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                
                # Pinta toda a linha
                for col_num in range(1, len(df_csv_original.columns) + 1):
                    cell = worksheet.cell(row=excel_row, column=col_num)
                    cell.fill = fill
                    cell.border = border
        
        # ====== APLICA AUTOFILTER NOS CABEÇALHOS ======
        ultima_linha_dados = len(df_csv_original) + 1
        ultima_coluna = len(df_csv_original.columns)
        range_filtro = f"A1:{get_column_letter(ultima_coluna)}{ultima_linha_dados}"
        worksheet.auto_filter.ref = range_filtro
        
        # Ajusta largura das colunas restantes automaticamente
        for col_num, column in enumerate(df_csv_original.columns, 1):
            col_letter = get_column_letter(col_num)
            # Pula colunas que já foram ajustadas manualmente
            if col_letter in ['I', 'L', 'M']:
                continue
            # Pula colunas que estão ocultas
            if col_letter in colunas_para_ocultar:
                continue
                
            max_length = len(str(column))
            for cell in worksheet[col_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # ====== POSICIONAMENTO DA TABELA DE RESUMO ======
        # Adiciona tabela de resumo (4 linhas abaixo dos dados para acomodar a fórmula)
        linha_inicio_resumo = len(df_csv_original) + 5
        
        # ====== ADICIONA FÓRMULA SUBTOTAL (QUE RESPEITA FILTROS) ======
        linha_formula = len(df_csv_original) + 3
        primeira_linha_dados = 2
        ultima_linha_dados = len(df_csv_original) + 1
        col_letter_valor = get_column_letter(coluna_valor_liberar)
        
        # CORREÇÃO: Usa SUBTOTAL(9,...) que soma apenas células visíveis (respeita filtros)
        # 109 = função SUM que ignora linhas ocultas por filtro
        formula = f"=SUBTOTAL(9,{col_letter_valor}{primeira_linha_dados}:{col_letter_valor}{ultima_linha_dados})"
        
        # Adiciona rótulo "Total" na coluna anterior
        col_rotulo = coluna_valor_liberar - 1
        if col_rotulo >= 1:
            worksheet.cell(row=linha_formula, column=col_rotulo, value="Subtotal")
            worksheet.cell(row=linha_formula, column=col_rotulo).font = Font(bold=True)
        
        # Adiciona a fórmula na coluna de valor
        cell_formula = worksheet.cell(row=linha_formula, column=coluna_valor_liberar)
        cell_formula.value = formula
        cell_formula.number_format = '#,##0.00'
        cell_formula.font = Font(bold=True)
        cell_formula.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Título do resumo - centralizado acima da tabela
        titulo_col = coluna_valor_liberar - 1  # Título começa uma coluna antes para centralização
        if titulo_col < 1:
            titulo_col = 1
        
        worksheet.cell(row=linha_inicio_resumo, column=titulo_col, value="RESUMO - VALOR A LIBERAR POR MOEDA")
        titulo_cell = worksheet.cell(row=linha_inicio_resumo, column=titulo_col)
        titulo_cell.font = Font(bold=True, size=12)
        
        linha_inicio_resumo += 2
        
        # ====== ESCREVE CABEÇALHOS DO RESUMO ALINHADOS ======
        colunas_resumo = ["Moeda", "Valor a Liberar", "Cotação", "Data da Cotação", "Valor em BRL"]
        
        # Posiciona "Moeda" uma coluna antes do "Valor a Liberar"
        col_inicio = coluna_valor_liberar - 1
        if col_inicio < 1:
            col_inicio = 1
        
        for col_num, col_name in enumerate(colunas_resumo, col_inicio):
            cell = worksheet.cell(row=linha_inicio_resumo, column=col_num, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # ====== ESCREVE DADOS DO RESUMO ALINHADOS ======
        for idx, row in df_resumo.iterrows():
            excel_row = linha_inicio_resumo + idx + 1
            
            # Moeda (uma coluna antes do Valor a Liberar)
            cell_moeda = worksheet.cell(row=excel_row, column=col_inicio, value=row["Moeda"])
            
            # Pinta a célula da moeda com a cor correspondente (exceto TOTAL)
            if row["Moeda"] != "TOTAL":
                cor_moeda = CORES_MOEDAS.get(row["Moeda"], "FFFFFF")
                fill_moeda = PatternFill(start_color=cor_moeda, end_color=cor_moeda, fill_type="solid")
                cell_moeda.fill = fill_moeda
            
            if row["Moeda"] != "TOTAL":
                # Valor a Liberar (numérico) - na coluna alinhada
                col_valor = col_inicio + 1
                cell_valor_liberar = worksheet.cell(row=excel_row, column=col_valor, value=row["Valor a Liberar"])
                cell_valor_liberar.number_format = '#,##0.00'
                
                # Cotação
                col_cotacao = col_inicio + 2
                if isinstance(row["Cotação"], (int, float)):
                    cell_cotacao = worksheet.cell(row=excel_row, column=col_cotacao, value=row["Cotação"])
                    cell_cotacao.number_format = '#,##0.00000'
                else:
                    worksheet.cell(row=excel_row, column=col_cotacao, value=row["Cotação"])
                
                # Data da Cotação
                col_data = col_inicio + 3
                worksheet.cell(row=excel_row, column=col_data, value=row["Data da Cotação"])
                
                # Valor em BRL - COM MÁSCARA DE REAIS (apenas se não for SDR)
                col_brl = col_inicio + 4
                if row["Moeda"] == "Direito Especial - SDR":
                    worksheet.cell(row=excel_row, column=col_brl, value="-")
                else:
                    cell_brl = worksheet.cell(row=excel_row, column=col_brl, value=row["Valor em BRL"])
                    cell_brl.number_format = '"R$" #,##0.00'
                
            else:
                # Linha TOTAL
                col_valor = col_inicio + 1
                col_cotacao = col_inicio + 2
                col_data = col_inicio + 3
                col_brl = col_inicio + 4
                
                worksheet.cell(row=excel_row, column=col_valor, value="-")
                worksheet.cell(row=excel_row, column=col_cotacao, value="-")
                worksheet.cell(row=excel_row, column=col_data, value="-")
                cell_brl = worksheet.cell(row=excel_row, column=col_brl, value=row["Valor em BRL"])
                cell_brl.number_format = '"R$" #,##0.00'
                
                # Formatação especial para linha TOTAL
                total_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                for col in range(col_inicio, col_inicio + 5):
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
            
            # Aplica bordas a todas as células do resumo
            for col in range(col_inicio, col_inicio + 5):
                worksheet.cell(row=excel_row, column=col).border = border
        
        # Ajusta largura das colunas do resumo
        for col_num in range(col_inicio, col_inicio + 5):
            col_letter = get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = 20
    
    buffer.seek(0)
    logger.info("Arquivo Excel completo gerado com sucesso")
    return buffer


# ====== Interface Streamlit ======
st.set_page_config(
    page_title="Resumo de Dívidas CDP",
    page_icon="💱",
    layout="centered"
)

st.title("💱 Gerar resumo do valor a liberar das dívidas no CDP")

# Upload do arquivo
uploaded_file = st.file_uploader(
    "Faça upload do arquivo [...]02-dividas.csv exportado do CDP do EF no SADIPEM",
    type="csv",
    help="Selecione o arquivo [...]02-dividas.csv"
)

if uploaded_file:
    # Verifica tamanho do arquivo
    file_size_mb = uploaded_file.size / (1024 * 1024)
    if file_size_mb > MAX_ARQUIVO_MB:
        st.error(f"❌ Arquivo muito grande ({file_size_mb:.1f}MB). Tamanho máximo: {MAX_ARQUIVO_MB}MB")
        st.stop()
    
    try:
        with st.spinner('🔄 Processando arquivo...'):
            # Lê o arquivo CSV
            logger.info(f"Lendo arquivo CSV: {uploaded_file.name} ({file_size_mb:.2f}MB)")
            df_csv = pd.read_csv(
                uploaded_file, 
                sep=";", 
                encoding="cp1252", 
                thousands=".", 
                decimal=","
            )
            
            # Valida estrutura
            validar_csv(df_csv)
            
            # Processa os dados
            df_resumo = processar_csv(df_csv)
            
            # Extrai registros detalhados
            df_detalhes = extrair_registros_detalhados(df_csv)
            
            # Verifica se encontrou registros
            if df_resumo is None:
                st.warning("⚠️ **Nenhum registro encontrado que atenda aos critérios**")
                st.info("""
                O arquivo foi processado, mas não foram encontrados registros que atendam aos seguintes critérios:
                
                - **Tipo de dívida**: "Empréstimo ou financiamento"
                - **Situação da dívida**: "Vigente"
                - **Valor a liberar**: Maior que zero
                """)
                st.stop()
            
            # Formata para exibição
            df_vis = formatar_para_exibicao(df_resumo)
            df_detalhes_vis = formatar_detalhes_para_exibicao(df_detalhes) if df_detalhes is not None else None
        
        # Exibe resultado
        st.success("✅ Processamento concluído com sucesso!")
        
        st.subheader("📊 Valor a Liberar por Moeda e Total em Reais")
        
        # Exibe tabela HTML customizada
        html_tabela = gerar_html_tabela(df_vis)
        st.markdown(html_tabela, unsafe_allow_html=True)
        
        # Botão de download
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            buffer_excel = gerar_excel_completo(df_csv, df_resumo)
            st.download_button(
                label="📥 Download em Excel (.xlsx)",
                data=buffer_excel.getvalue(),
                file_name=f"resumo_dividas_valor_liberar_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        # Informações adicionais
        st.caption(f"📅 Processado em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}")
        st.caption(f"📄 {len(df_resumo) - 1} moeda(s) utilizada(s)")
        
        # ====== REGISTROS DETALHADOS ======
        if df_detalhes_vis is not None:
            st.divider()
            st.subheader("📋 Registros de dívida com valor a liberar", 
                         help="💡 Para melhor visualização da tabela, clique nos 3 pontos (⋮) no canto superior direito, depois em settings e ative 'Wide mode'")
            
            # Exibe tabela HTML customizada de detalhes
            html_tabela_detalhes = gerar_html_tabela_detalhes(df_detalhes_vis)
            st.markdown(html_tabela_detalhes, unsafe_allow_html=True)
            
            st.caption(f"📊 Total de {len(df_detalhes_vis)} registro(s)")
        
    except ValueError as ve:
        st.error(f"❌ Erro de validação: {ve}")
        logger.error(f"Erro de validação: {ve}")
        
    except Exception as e:
        st.error(f"❌ Erro ao processar o arquivo: {e}")
        logger.exception("Erro durante processamento")
        st.info("💡 Verifique se o arquivo está no formato correto exportado do CDP")

# Informações e instruções
with st.expander("ℹ️ Instruções de Uso"):
    st.markdown("""
    ### Como usar este aplicativo:
    
    1. **Fazer upload**: Faça o upload do arquivo `[...]02-dividas.csv` exportado do CDP do Ente Federativo no SADIPEM
    
    2. **Processamento automático**: O sistema irá:
       - Filtrar apenas dívidas do tipo "Empréstimo ou financiamento"
       - Considerar somente dívidas com situação "Vigente"
       - Incluir apenas valores a liberar maiores que zero
       - Agrupar os valores por moeda
       - Converter para Real (BRL) usando cotações oficiais PTAX do Banco Central
    
    3. **Download**: Após o processamento, faça o download da planilha Excel com:
       - Todos os registros do CSV original (linhas coloridas por moeda quando atendem aos critérios)
       - Tabela de resumo abaixo com totais por moeda e conversão para BRL
       - Layout amigável: estrutura da planilha facilita confrontar os registros originais com os subtotais por moeda

    ---
    
    ### Sobre as cotações:
    **Fonte**: Os valores em moeda estrangeira são convertidos para Real utilizando a cotação PTAX de venda do Banco Central do Brasil, referente ao fechamento do dia
    
    **Data da cotação**: A data da cotação é o último dia do RREO exigível (último dia do bimestre) na data corrente; ou data útil anterior caso caia em final de semana ou feriado
    
    📅 Datas das Cotações:

    | RREO exigível | Período de análise | Data da cotação* |
    |----------|-----------------|-------------------|
    | 1º Bimestre | 31/03 a 30/05 | **28/02** |
    | 2º Bimestre | 31/05 a 30/07 | **30/04** |
    | 3º Bimestre | 31/07 a 30/09 | **30/06** |
    | 4º Bimestre | 01/10 a 30/11 | **31/08** |
    | 5º Bimestre | 01/12 a 30/01 (ano seguinte) | **31/10** |
    | 6º Bimestre | 31/01 a 30/03 | **31/12** |

    **Ou dia útil anterior*
    
    **SDR**: Para Direitos Especiais de Saque (SDR), não há cotação disponível na API PTAX, portanto o valor não é convertido para BRL
    
    ---
    
    ### Moedas suportadas:
    - Real (BRL)
    - Dólar dos EUA (USD)
    - Euro (EUR)
    - Iene (JPY)

    """)

with st.expander("🔧 Informações Técnicas"):
    st.markdown("""
    ### Tecnologias utilizadas:
    - **Streamlit**: Interface web interativa
    - **Pandas**: Processamento e análise de dados
    - **BCB (python-bcb)**: Integração com API do Banco Central
    - **OpenPyXL**: Geração de arquivos Excel com formatação
    - **Python 3.x**: Linguagem de programação
    
    ### Critérios de filtragem:
    ```
    Tipo de dívida = "Empréstimo ou financiamento"
    Situação da dívida = "Vigente"
    Valor a liberar > 0
    ```
    
    ### Formatação do Excel:
    - **AutoFiltro**: Filtros automáticos em todas as colunas dos dados originais
    - **Fórmula SUBTOTAL**: Total geral dinâmico que se ajusta aos filtros aplicados
    - **Linhas coloridas**: Registros que atendem aos critérios são pintados por moeda
    - **Cores por moeda**: Verde (Real), Amarelo (Dólar), Azul (Euro), Laranja (SDR), Verde água (Iene)
    - **Legenda visual**: Células da coluna "Moeda" na tabela de resumo pintadas com as cores correspondentes
    - **Formato numérico**: Padrão brasileiro (#.##0,00)
    - **Formatação de moeda**: Coluna "Valor a liberar ou assumir (na moeda de contratação)" formatada com máscara de moeda genérica
    - **Resumo destacado**: Tabela de resumo com linha TOTAL em azul claro
    - **Máscara de Reais**: Valores em BRL formatados com "R$" no Excel
    - **Colunas ocultas**: A, D, E, G, H, J, K, O, W até AE
    
    ### Logs e cache:
    - Sistema de logs configurado para rastreabilidade
    - Cache de cotações por 8 horas (reduz chamadas à API)
    - Validações em múltiplas etapas do processamento
    """)

# Rodapé
st.divider()
st.markdown("""
<div style='text-align: center; color: #666; font-size: 12px;'>
    <p>Secretaria do Tesouro Nacional - STN</p>
</div>
""", unsafe_allow_html=True)
